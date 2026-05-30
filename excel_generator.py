import os
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter
from openpyxl.workbook.views import BookView

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY    = "1B2A4A"
WHITE   = "FFFFFF"
GOLD    = "F5A623"
LIGHT   = "F7F8FA"
MID     = "E8EBF0"
ACCENT  = "2563EB"   # bright blue for chart fill
ACCENT2 = "16A34A"   # green for salary present
DANGER  = "DC2626"   # red for salary missing

COLUMNS = {
    "indeed":   ["Title", "Company", "Location", "Salary", "Recruiter", "Phone", "URL"],
    "linkedin": ["Title", "Company", "Location", "Posted", "Recruiter", "Phone", "URL"],
}

KEY_MAP = {
    "indeed":   ["title", "company", "location", "salary", "recruiter_name", "recruiter_phone", "url"],
    "linkedin": ["title", "company", "location", "posted", "recruiter_name", "recruiter_phone", "url"],
}


# ── Public entry point ────────────────────────────────────────────────────────
def generate_excel(all_results: dict[str, list[dict]], scraper: str, config: dict) -> str:
    """
    all_results: {"indeed": [...], "linkedin": [...]}  (one or both keys)
    scraper:     "indeed" | "linkedin" | "both"
    Produces one workbook with 3 sheets per source (Dashboard, Jobs, Charts).
    """
    if not any(all_results.values()):
        raise ValueError("All scrapers returned no data — nothing to write.")

    out_cfg   = config.get("output", {})
    directory = out_cfg.get("directory", "output")
    os.makedirs(directory, exist_ok=True)

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = out_cfg.get("filename", "{scraper}_jobs_{date}.xlsx")
    filename = filename.replace("{scraper}", scraper).replace("{date}", date_str)
    filepath = os.path.join(directory, filename)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    wb.views = [BookView(windowWidth=28800, windowHeight=17460, xWindow=0, yWindow=0)]

    total_rows = 0
    for source_name, data in all_results.items():
        if not data:
            continue
        prefix = f"{source_name.capitalize()} - " if len(all_results) > 1 else ""
        _build_dashboard(wb, data, source_name, date_str, prefix=prefix)
        _build_jobs_sheet(wb, data, source_name, prefix=prefix)
        _build_charts_sheet(wb, data, source_name, prefix=prefix)
        total_rows += len(data)

    wb.save(filepath)
    print(f"[excel] Saved {total_rows} total rows → {filepath}")
    return filepath


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────
def _build_dashboard(wb: Workbook, data: list[dict], scraper: str, date_str: str, prefix: str = ""):
    ws = wb.create_sheet(f"{prefix}Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 28

    # Banner
    ws.merge_cells("B1:F2")
    banner = ws["B1"]
    banner.value = f"  {'Indeed' if scraper == 'indeed' else 'LinkedIn'}  Job Market Report"
    banner.font = Font(name="Calibri", size=20, bold=True, color=WHITE)
    banner.fill = PatternFill("solid", fgColor=NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18

    # Sub-header line
    ws.merge_cells("B3:F3")
    sub = ws["B3"]
    sub.value = f"Generated: {date_str}   |   Source: {scraper.capitalize()}"
    sub.font = Font(name="Calibri", size=10, color="888888")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 16

    ws.row_dimensions[4].height = 10  # spacer

    # KPI stats
    companies = [r.get("company", "") for r in data if r.get("company")]
    locations = [r.get("location", "") for r in data if r.get("location")]
    top_location    = Counter(locations).most_common(1)[0][0] if locations else "—"
    recruiters_with_name = sum(
        1 for r in data if r.get("recruiter_name", "Not listed") != "Not listed"
    )

    if scraper == "indeed":
        salary_listed = sum(1 for r in data if r.get("salary", "Not listed") != "Not listed")
        salary_pct    = f"{salary_listed / len(data) * 100:.0f}%" if data else "0%"
        kpis = [
            ("Total Jobs",       str(len(data))),
            ("Unique Companies", str(len(set(companies)))),
            ("Salary Coverage",  salary_pct),
            ("Recruiter Data",   f"{recruiters_with_name} jobs"),
        ]
    else:
        kpis = [
            ("Total Jobs",       str(len(data))),
            ("Unique Companies", str(len(set(companies)))),
            ("Top Location",     top_location),
            ("Recruiter Data",   f"{recruiters_with_name} jobs"),
        ]

    for i, (label, value) in enumerate(kpis):
        col = chr(ord("C") + i)   # C, D, E, F
        row_label = 5
        row_value = 6

        label_cell = ws[f"{col}{row_label}"]
        label_cell.value = label
        label_cell.font  = Font(name="Calibri", size=9, color="888888", bold=False)
        label_cell.alignment = Alignment(horizontal="center")

        val_cell = ws[f"{col}{row_value}"]
        val_cell.value = value
        val_cell.font  = Font(name="Calibri", size=22, bold=True, color=NAVY)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill  = PatternFill("solid", fgColor=LIGHT)

        ws.row_dimensions[row_label].height = 14
        ws.row_dimensions[row_value].height = 34

    # Divider
    ws.row_dimensions[7].height = 8
    ws.merge_cells("B8:F8")
    div = ws["B8"]
    div.fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[8].height = 3

    # Top 5 companies table
    ws.row_dimensions[9].height = 10  # spacer
    _dash_section_header(ws, "B10", "Top 5 Companies by Listings")
    ws.row_dimensions[10].height = 18

    top5 = Counter(companies).most_common(5)
    for idx, (name, count) in enumerate(top5, start=11):
        ws[f"B{idx}"] = name
        ws[f"B{idx}"].font = Font(name="Calibri", size=10, color=NAVY)
        ws[f"C{idx}"] = count
        ws[f"C{idx}"].font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
        ws.row_dimensions[idx].height = 14

    # Top 5 locations table (column E–F)
    _dash_section_header(ws, "E10", "Top 5 Locations")
    top5_loc = Counter(locations).most_common(5)
    for idx, (loc, count) in enumerate(top5_loc, start=11):
        ws[f"E{idx}"] = loc
        ws[f"E{idx}"].font = Font(name="Calibri", size=10, color=NAVY)
        ws[f"F{idx}"] = count
        ws[f"F{idx}"].font = Font(name="Calibri", size=10, bold=True, color=ACCENT)


def _dash_section_header(ws, cell_ref: str, title: str):
    c = ws[cell_ref]
    c.value = title
    c.font  = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ── Sheet 2: Jobs ─────────────────────────────────────────────────────────────
def _build_jobs_sheet(wb: Workbook, data: list[dict], scraper: str, prefix: str = ""):
    ws = wb.create_sheet(f"{prefix}Jobs")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A2"

    columns = COLUMNS.get(scraper, list(data[0].keys()))
    keys    = KEY_MAP.get(scraper, [c.lower() for c in columns])

    # Header row
    header_fill   = PatternFill("solid", fgColor=NAVY)
    header_font   = Font(name="Calibri", size=11, bold=True, color=WHITE)
    alt_fill      = PatternFill("solid", fgColor=LIGHT)
    normal_fill   = PatternFill("solid", fgColor=WHITE)
    body_font     = Font(name="Calibri", size=10, color=NAVY)
    url_font      = Font(name="Calibri", size=10, color=ACCENT, underline="single")
    thin_border   = Border(bottom=Side(style="thin", color=MID))

    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Auto-filter on header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Data rows
    for row_idx, row in enumerate(data, start=2):
        fill = alt_fill if row_idx % 2 == 0 else normal_fill
        for col_idx, key in enumerate(keys, start=1):
            value = row.get(key, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = thin_border
            if key == "url" and value:
                cell.value     = "View Job"
                cell.hyperlink = value
                cell.font      = url_font
            else:
                cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "title"))
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    col_widths = {
        "title": 48, "company": 24, "location": 22, "salary": 18,
        "posted": 16, "recruiter_name": 22, "recruiter_phone": 16, "url": 12,
    }
    for col_idx, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 20)


# ── Sheet 3: Charts ───────────────────────────────────────────────────────────
def _build_charts_sheet(wb: Workbook, data: list[dict], scraper: str, prefix: str = ""):
    ws = wb.create_sheet(f"{prefix}Charts")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 80

    companies = [r.get("company", "Unknown") or "Unknown" for r in data]
    locations = [r.get("location", "Unknown") or "Unknown" for r in data]

    top_companies = Counter(companies).most_common(10)
    top_locations = Counter(locations).most_common(10)

    # Write data tables for charts (hidden helper data, cols M onwards)
    # Companies table
    ws["M1"] = "Company"
    ws["N1"] = "Count"
    for i, (name, count) in enumerate(top_companies, start=2):
        ws[f"M{i}"] = name
        ws[f"N{i}"] = count

    # Locations table
    ws["P1"] = "Location"
    ws["Q1"] = "Count"
    for i, (loc, count) in enumerate(top_locations, start=2):
        ws[f"P{i}"] = loc
        ws[f"Q{i}"] = count

    # Salary / recency for pie chart (col S–T)
    if scraper == "indeed":
        salary_yes = sum(1 for r in data if r.get("salary", "Not listed") not in ("Not listed", "", None))
        salary_no  = len(data) - salary_yes
        ws["S1"] = "Salary Status"
        ws["T1"] = "Count"
        ws["S2"] = "Salary Listed"
        ws["T2"] = salary_yes
        ws["S3"] = "Not Listed"
        ws["T3"] = salary_no
        pie_rows = 3
        pie_title = "Salary Coverage"
    else:
        from datetime import date
        today = date.today()
        buckets = {"Today": 0, "This Week": 0, "This Month": 0, "Older": 0}
        for r in data:
            posted = r.get("posted", "")
            if "hour" in posted or "minute" in posted or "just now" in posted.lower():
                buckets["Today"] += 1
            elif "day" in posted and int("".join(filter(str.isdigit, posted) or "99")) <= 7:
                buckets["This Week"] += 1
            elif "week" in posted:
                buckets["This Month"] += 1
            else:
                buckets["Older"] += 1
        ws["S1"] = "Posted"
        ws["T1"] = "Count"
        for i, (k, v) in enumerate(buckets.items(), start=2):
            ws[f"S{i}"] = k
            ws[f"T{i}"] = v
        pie_rows = len(buckets) + 1
        pie_title = "Jobs by Recency"

    # ── Chart 1: Top Companies (Bar) ──
    bar1 = BarChart()
    bar1.type    = "bar"
    bar1.title   = "Top 10 Companies by Listings"
    bar1.y_axis.title = "Company"
    bar1.x_axis.title = "Job Count"
    bar1.style   = 10
    bar1.width   = 24
    bar1.height  = 14
    bar1.grouping = "clustered"

    data_ref = Reference(ws, min_col=14, min_row=1, max_row=len(top_companies) + 1)
    cats_ref = Reference(ws, min_col=13, min_row=2, max_row=len(top_companies) + 1)
    bar1.add_data(data_ref, titles_from_data=True)
    bar1.set_categories(cats_ref)
    bar1.series[0].graphicalProperties.solidFill = ACCENT
    bar1.series[0].graphicalProperties.line.solidFill = ACCENT
    ws.add_chart(bar1, "B2")

    # ── Chart 2: Top Locations (Bar) ──
    bar2 = BarChart()
    bar2.type    = "bar"
    bar2.title   = "Top 10 Locations by Listings"
    bar2.y_axis.title = "Location"
    bar2.x_axis.title = "Job Count"
    bar2.style   = 10
    bar2.width   = 24
    bar2.height  = 14
    bar2.grouping = "clustered"

    data_ref2 = Reference(ws, min_col=17, min_row=1, max_row=len(top_locations) + 1)
    cats_ref2 = Reference(ws, min_col=16, min_row=2, max_row=len(top_locations) + 1)
    bar2.add_data(data_ref2, titles_from_data=True)
    bar2.set_categories(cats_ref2)
    bar2.series[0].graphicalProperties.solidFill = GOLD
    bar2.series[0].graphicalProperties.line.solidFill = GOLD
    ws.add_chart(bar2, "B28")

    # ── Chart 3: Pie chart ──
    pie = PieChart()
    pie.title  = pie_title
    pie.style  = 10
    pie.width  = 18
    pie.height = 14

    pie_data = Reference(ws, min_col=20, min_row=1, max_row=pie_rows)
    pie_cats = Reference(ws, min_col=19, min_row=2, max_row=pie_rows)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "L2")

    # Section labels
    _chart_label(ws, "B1", "Companies")
    _chart_label(ws, "B27", "Locations")
    _chart_label(ws, "L1", pie_title)


def _chart_label(ws, cell_ref: str, text: str):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(name="Calibri", size=12, bold=True, color=NAVY)
